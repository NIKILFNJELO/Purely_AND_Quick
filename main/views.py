import json
import re
import logging

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST, require_http_methods
from django.views.generic import TemplateView
from django.http import JsonResponse

from .models import Profile, Order, Review, CategoryConfig

logger = logging.getLogger(__name__)


class HomePageView(TemplateView):
    template_name = 'main/about_us.html'

class ProposalsView(TemplateView):
    template_name = 'main/proposals.html'


# ── Дефолтні дані категорій ──────────────────────────────────────────────────
_DEFAULT_BRANCHES = ["Нові будинки", "Центр", "Салтівка", "Мотель", "П'ятихатки", "ХТЗ"]

_CATEGORY_DEFAULTS = {
    "Побутовий одяг": {
        "image_url": "https://organic-store.in.ua/images/beauty/beauty-1458.jpg",
        "items": [
            {"name": "Куртка / пальто",  "price": 280},
            {"name": "Светр / кофта",    "price": 150},
            {"name": "Штани / джинси",   "price": 130},
            {"name": "Сукня / спідниця", "price": 160},
            {"name": "Сорочка / блуза",  "price": 100},
        ],
        "services": [
            {"name": "Хімчистка",                 "price": 150},
            {"name": "Прасування",                "price":  60},
            {"name": "Виведення плям",            "price":  80},
            {"name": "Ремонт тканини",            "price": 120},
            {"name": "Водовідштовх. просочення",  "price": 100},
            {"name": "Антистат. обробка",         "price":  70},
        ],
    },
    "Спортивний одяг": {
        "image_url": "https://images.prom.ua/4243735384_w640_h640_muzhskoj-sportivnyj-kostyum.jpg",
        "items": [
            {"name": "Спортивний костюм", "price": 200},
            {"name": "Футболка / майка",  "price":  80},
            {"name": "Спортивні штани",   "price": 110},
            {"name": "Термобілизна",      "price": 140},
            {"name": "Кросівки / кеди",   "price": 180},
        ],
        "services": [
            {"name": "Хімчистка",            "price": 120},
            {"name": "Виведення поту",       "price":  90},
            {"name": "Дезодорація",          "price":  70},
            {"name": "Виведення трав. плям", "price":  85},
            {"name": "Відновлення кольору",  "price": 150},
            {"name": "Антибакт. обробка",    "price":  80},
        ],
    },
    "Робочий одяг": {
        "image_url": "https://togtrade.com/image/catalog/jak-organizuvati-reguljarnu-zaminu-specodjagu-na-virobnictvi-2.png",
        "items": [
            {"name": "Комбінезон / роба", "price": 250},
            {"name": "Захисна куртка",    "price": 220},
            {"name": "Робочі штани",      "price": 150},
            {"name": "Спецодяг медика",   "price": 180},
            {"name": "Фартух / жилет",    "price": 100},
        ],
        "services": [
            {"name": "Промислова хімчистка",    "price": 200},
            {"name": "Виведення мастила",       "price": 130},
            {"name": "Виведення фарби",         "price": 150},
            {"name": "Дезінфекція",             "price": 100},
            {"name": "Ремонт швів",             "price":  90},
            {"name": "Вогнезахист. просочення", "price": 180},
        ],
    },
    "Фірмовий одяг": {
        "image_url": "https://images.unian.net/photos/2022_09/thumb_files/400_0_1662204765-9857.jpg?r=814461",
        "items": [
            {"name": "Діловий костюм",  "price": 350},
            {"name": "Сорочка / блуза", "price": 120},
            {"name": "Краватка / шарф", "price":  80},
            {"name": "Піджак / жакет",  "price": 260},
            {"name": "Форменний одяг",  "price": 200},
        ],
        "services": [
            {"name": "Делікатна хімчистка", "price": 200},
            {"name": "Прасування з парою",  "price":  90},
            {"name": "Виведення плям",      "price": 100},
            {"name": "Відновлення форми",   "price": 160},
            {"name": "Чищення підкладки",   "price": 120},
            {"name": "Нашивка емблеми",     "price":  70},
        ],
    },
}


def _get_or_create_category(cat_name):
    if cat_name in _CATEGORY_DEFAULTS:
        defaults = _CATEGORY_DEFAULTS[cat_name]
        obj, _ = CategoryConfig.objects.get_or_create(
            category=cat_name,
            defaults={
                "items":     defaults["items"],
                "services":  defaults["services"],
                "branches":  _DEFAULT_BRANCHES,
                "image_url": defaults.get("image_url", ""),
            }
        )
    else:
        # Нова категорія додана менеджером — просто get або create порожнім
        obj, _ = CategoryConfig.objects.get_or_create(
            category=cat_name,
            defaults={"items": [], "services": [], "branches": _DEFAULT_BRANCHES}
        )
    return obj


# ── Views ────────────────────────────────────────────────────────────────────

def catalog(request):
    has_discount = False
    if request.user.is_authenticated:
        profile, _ = Profile.objects.get_or_create(user=request.user, defaults={"phone": ""})
        has_discount = profile.paid_orders_count >= 3

    # Спочатку ensure всі дефолтні категорії є в БД
    for cat_name in _CATEGORY_DEFAULTS:
        _get_or_create_category(cat_name)
    # Підтягуємо ВСІ категорії з БД (включаючи додані менеджером)
    categories = []
    for obj in CategoryConfig.objects.all():
        obj.items_json    = json.dumps(obj.items    or [], ensure_ascii=False)
        obj.services_json = json.dumps(obj.services or [], ensure_ascii=False)
        obj.branches_json = json.dumps(obj.branches or [], ensure_ascii=False)
        categories.append(obj)

    return render(request, "main/catalog.html", {
        "has_discount":       has_discount,
        "user_authenticated": request.user.is_authenticated,
        "categories":         categories,
    })


@require_POST
@login_required
def save_category_config(request):
    if not request.user.is_staff:
        return JsonResponse({"error": "Forbidden"}, status=403)
    try:
        data     = json.loads(request.body)
        category = data.get("category", "").strip()
        if not category:
            return JsonResponse({"error": "category required"}, status=400)
        obj, _ = CategoryConfig.objects.get_or_create(category=category)
        if "items"     in data: obj.items     = data["items"]
        if "services"  in data: obj.services  = data["services"]
        if "branches"  in data: obj.branches  = data["branches"]
        if "image_url" in data: obj.image_url = data["image_url"]
        obj.save()
        return JsonResponse({"status": "ok"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)



@require_POST
@login_required
def delete_category_config(request):
    if not request.user.is_staff:
        return JsonResponse({"error": "Forbidden"}, status=403)
    try:
        data = json.loads(request.body)
        category = data.get("category", "").strip()
        if not category:
            return JsonResponse({"error": "category required"}, status=400)
        deleted, _ = CategoryConfig.objects.filter(category=category).delete()
        if deleted:
            return JsonResponse({"status": "ok"})
        return JsonResponse({"error": "Категорію не знайдено"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)

@login_required
def profile(request):
    if request.method == "POST" and request.FILES.get("avatar"):
        request.user.profile.avatar = request.FILES["avatar"]
        request.user.profile.save()
        return redirect("profile")

    orders = Order.objects.filter(user=request.user).order_by('id')
    for i, order in enumerate(orders, start=1):
        order.user_number = i

    orders_until_discount = max(0, 3 - request.user.profile.paid_orders_count)

    # Ensure дефолтні є в БД
    for cat_name in _CATEGORY_DEFAULTS:
        _get_or_create_category(cat_name)
    # catalog_data — всі категорії з БД для JS (edit modal)
    catalog_data = {}
    for obj in CategoryConfig.objects.all():
        catalog_data[obj.category] = {
            "items":    obj.items    or [],
            "services": obj.services or [],
            "branches": obj.branches or [],
        }

    return render(request, "main/profile.html", {
        "orders":                orders,
        "orders_until_discount": orders_until_discount,
        "catalog_json":          json.dumps(catalog_data, ensure_ascii=False),
    })


@login_required
def order_list(request):
    orders = Order.objects.filter(user=request.user).order_by('id')
    for i, order in enumerate(orders, start=1):
        order.user_number = i
    return render(request, "main/order_list.html", {"orders": orders})


@login_required
@require_POST
def create_order(request):
    item_name    = request.POST.get("item_name", "").strip()
    branch       = request.POST.get("branch", "").strip()
    desired_date = request.POST.get("desired_date", "").strip()
    defect       = request.POST.get("defect_description", "").strip()
    urgent       = request.POST.get("urgent") == "True"

    # ✅ getlist — отримує ВСІ вибрані чекбокси
    service_list = request.POST.getlist("service_type")
    service_type = ", ".join(s.strip() for s in service_list if s.strip())

    try:
        total_price = int(request.POST.get("total_price", "0"))
    except (ValueError, TypeError):
        total_price = 0

    if not item_name or not service_type or not branch or not desired_date:
        messages.error(request, "Заповніть всі обов'язкові поля")
        return redirect("catalog")

    Order.objects.create(
        user=request.user,
        item_name=item_name,
        service_type=service_type,
        branch=branch,
        desired_date=desired_date,
        defect_description=defect,
        urgent=urgent,
        total_price=total_price if total_price > 0 else 0,
    )
    messages.success(request, "Замовлення успішно оформлено!")
    return redirect("profile")


@login_required
@require_http_methods(["GET", "POST"])
def edit_order(request, order_id):
    # ✅ ЄДИНА версія edit_order — без дублювання
    order = get_object_or_404(Order, id=order_id, user=request.user)

    if order.is_paid:
        messages.error(request, "Не можна редагувати сплачене замовлення")
        return redirect("profile")

    if request.method == "POST":
        item_name    = request.POST.get("item_name", "").strip()
        branch       = request.POST.get("branch", "").strip()
        desired_date = request.POST.get("desired_date", "").strip()
        defect       = request.POST.get("defect_description", "").strip()
        urgent       = request.POST.get("urgent") == "True"

        # З profile.html приходить hidden input з готовим рядком "Послуга1, Послуга2"
        # зібраним JS перед submit. getlist поверне список з одного елемента.
        service_raw  = request.POST.get("service_type", "").strip()
        service_type = service_raw  # вже готовий рядок

        try:
            total_price = int(request.POST.get("total_price", "0"))
        except (ValueError, TypeError):
            total_price = 0

        if not item_name or not service_type or not branch or not desired_date:
            messages.error(request, "Заповніть всі обов'язкові поля")
            return redirect("profile")

        order.item_name          = item_name
        order.service_type       = service_type
        order.branch             = branch
        order.desired_date       = desired_date
        order.defect_description = defect
        order.urgent             = urgent
        if total_price > 0:
            order.total_price = total_price
        order.save()

        messages.success(request, "Замовлення оновлено!")
        return redirect("profile")

    return redirect("profile")


@require_POST
@login_required
def pay_order(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    order.is_paid = True
    order.save()
    profile = request.user.profile
    profile.paid_orders_count += 1
    if profile.paid_orders_count >= 3:
        profile.loyalty_status = "Постійний клієнт"
    profile.save()
    return JsonResponse({'status': 'ok'})


@require_POST
@login_required
def delete_order(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    order.delete()
    return JsonResponse({'status': 'deleted'})


def order_detail(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    return render(request, "main/order_detail.html", {"order": order})



# ── Manager dashboard ────────────────────────────────────────────────────────

@login_required
def manager_dashboard(request):
    from django.db.models import Sum, Count, Q
    from django.utils import timezone
    from datetime import date, timedelta
    from collections import Counter

    if not request.user.is_staff:
        return redirect("profile")

    # ── Date filter ──
    date_from_str = request.GET.get("from", "")
    date_to_str   = request.GET.get("to", "")
    today = date.today()

    try:
        date_from = date.fromisoformat(date_from_str) if date_from_str else None
    except ValueError:
        date_from = None
    try:
        date_to = date.fromisoformat(date_to_str) if date_to_str else None
    except ValueError:
        date_to = None

    qs = Order.objects.all()
    if date_from:
        qs = qs.filter(order_date__gte=date_from)
    if date_to:
        qs = qs.filter(order_date__lte=date_to)

    # ── KPI ──
    total_orders  = qs.count()
    paid_orders   = qs.filter(is_paid=True).count()
    unpaid_orders = qs.filter(is_paid=False).count()
    urgent_orders = qs.filter(is_paid=False, urgent=True).count()

    paid_qs = qs.filter(is_paid=True)
    total_revenue = int(paid_qs.aggregate(s=Sum("total_price"))["s"] or 0)
    avg_check     = int(total_revenue / paid_orders) if paid_orders else 0

    total_clients = User.objects.filter(is_staff=False).count()
    loyal_clients = Profile.objects.filter(paid_orders_count__gte=3).count()

    # ── Branch stats ──
    BRANCH_COLORS_MAP = [
        "#d4a574","#3dba72","#5ca8e0","#ffb347",
        "#a87adb","#e05c6a","#4ecdc4","#f7dc6f",
    ]
    # Філії з БД (CategoryConfig) — об'єднуємо всі унікальні
    all_configs = {obj.category: obj for obj in CategoryConfig.objects.all()}
    db_branches = []
    seen = set()
    for obj in all_configs.values():
        for b in (obj.branches or []):
            if b not in seen:
                db_branches.append(b)
                seen.add(b)
    # Якщо в БД ще нічого немає — fallback на дефолт
    branch_names = db_branches if db_branches else _DEFAULT_BRANCHES
    branch_stats = []
    max_orders = 1

    raw_branch = (
        qs.values("branch")
          .annotate(
              orders=Count("id"),
              revenue=Sum("total_price", filter=Q(is_paid=True)),
              paid=Count("id", filter=Q(is_paid=True)),
          )
          .order_by("-orders")
    )
    branch_map = {b["branch"]: b for b in raw_branch}
    if raw_branch:
        max_orders = max((b["orders"] for b in raw_branch), default=1)

    for i, bname in enumerate(branch_names):
        b = branch_map.get(bname, {"orders": 0, "revenue": 0, "paid": 0})
        branch_stats.append({
            "branch":  bname,
            "orders":  b["orders"],
            "revenue": int(b["revenue"] or 0),
            "paid":    b["paid"],
            "pct":     round(b["orders"] / max_orders * 100) if max_orders else 0,
            "color":   BRANCH_COLORS_MAP[i % len(BRANCH_COLORS_MAP)],
        })

    # ── Top services ──
    service_counter = Counter()
    for svc_str in qs.values_list("service_type", flat=True):
        for s in svc_str.split(","):
            s = s.strip()
            if s and s.lower() != "on":
                service_counter[s] += 1

    top_services = [{"name": k, "count": v} for k, v in service_counter.most_common(7)]

    # ── Clients ──
    clients_raw = (
        User.objects.filter(is_staff=False)
            .annotate(
                order_count=Count("order"),
                paid_count=Count("order", filter=Q(order__is_paid=True)),
                total_spent=Sum("order__total_price", filter=Q(order__is_paid=True)),
            )
            .order_by("-paid_count")[:30]
    )
    clients = []
    for c in clients_raw:
        clients.append({
            "username":    c.username,
            "order_count": c.order_count,
            "paid_count":  c.paid_count,
            "total_spent": int(c.total_spent or 0),
            "is_loyal":    getattr(getattr(c, "profile", None), "paid_orders_count", 0) >= 3,
        })

    # ── Category guess helper — з БД ──
    ITEM_TO_CAT = {}
    for cat_name, obj in all_configs.items():
        for item in (obj.items or []):
            ITEM_TO_CAT[item["name"]] = cat_name
    # Fallback на дефолти якщо БД порожня
    if not ITEM_TO_CAT:
        for cat_name, cat_data in _CATEGORY_DEFAULTS.items():
            for item in cat_data["items"]:
                ITEM_TO_CAT[item["name"]] = cat_name

    all_orders = list(qs.select_related("user").order_by("-id"))
    for o in all_orders:
        o.category_guess = ITEM_TO_CAT.get(o.item_name, "")

    # ── Unique branches for filter dropdown ──
    branches = sorted(set(o.branch for o in all_orders if o.branch))

    # ── Catalog summary з БД для відображення в дашборді ──
    catalog_summary = []
    for cat_name in _CATEGORY_DEFAULTS:
        obj = all_configs.get(cat_name)
        if obj:
            catalog_summary.append({
                "category": obj.category,
                "items":    obj.items or [],
                "services": obj.services or [],
                "branches": obj.branches or [],
            })
        else:
            d = _CATEGORY_DEFAULTS[cat_name]
            catalog_summary.append({
                "category": cat_name,
                "items":    d["items"],
                "services": d["services"],
                "branches": _DEFAULT_BRANCHES,
            })

    return render(request, "main/manager.html", {
        "total_revenue":  total_revenue,
        "total_orders":   total_orders,
        "paid_orders":    paid_orders,
        "unpaid_orders":  unpaid_orders,
        "urgent_orders":  urgent_orders,
        "avg_check":      avg_check,
        "total_clients":  total_clients,
        "loyal_clients":  loyal_clients,
        "branch_stats":   branch_stats,
        "top_services":   top_services,
        "clients":        clients,
        "all_orders":     all_orders,
        "branches":       branches,
        "date_from":      date_from_str,
        "date_to":        date_to_str,
        "catalog_summary": catalog_summary,
    })


# ── Excel Export ─────────────────────────────────────────────────────────────

@login_required
def export_orders_excel(request):
    """Формує .xlsx звіт по замовленнях для менеджера."""
    import io
    from datetime import date
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.db.models import Sum, Count, Q

    if not request.user.is_staff:
        return redirect("profile")

    # ── Date filter ──
    date_from_str = request.GET.get("from", "")
    date_to_str   = request.GET.get("to", "")
    try:
        date_from = date.fromisoformat(date_from_str) if date_from_str else None
    except ValueError:
        date_from = None
    try:
        date_to = date.fromisoformat(date_to_str) if date_to_str else None
    except ValueError:
        date_to = None

    qs = Order.objects.select_related("user").order_by("branch", "id")
    if date_from:
        qs = qs.filter(order_date__gte=date_from)
    if date_to:
        qs = qs.filter(order_date__lte=date_to)

    orders = list(qs)

    # Беремо всі CategoryConfig з БД
    all_configs_export = {obj.category: obj for obj in CategoryConfig.objects.all()}
    # Унікальні філії з БД
    db_branches_export = []
    seen_b = set()
    for obj in all_configs_export.values():
        for b in (obj.branches or []):
            if b not in seen_b:
                db_branches_export.append(b)
                seen_b.add(b)
    if not db_branches_export:
        db_branches_export = _DEFAULT_BRANCHES

    # ── Workbook setup ──
    wb = Workbook()

    # ═══ STYLES ═══
    GOLD      = "D4A574"
    DARK      = "3B2A1A"
    LIGHT_BG  = "FDF6EE"
    HDR_BG    = "6B4423"
    ALT_ROW   = "FDF0E0"
    GREEN_BG  = "E8F8EE"
    RED_BG    = "FDE8EA"

    def hdr_font(size=11, bold=True, color="FFFFFF"):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def cell_font(size=10, bold=False, color=DARK):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def fill(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

    def border():
        s = Side(style="thin", color="D4B896")
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ═══════════════════════════════════════════════
    # SHEET 1 — Всі замовлення
    # ═══════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Всі замовлення"
    ws1.sheet_view.showGridLines = False

    # Title row
    ws1.merge_cells("A1:I1")
    title_cell = ws1["A1"]
    period = ""
    if date_from or date_to:
        period = f"  |  Період: {date_from_str or '—'} — {date_to_str or '—'}"
    title_cell.value = f"Purely&Quick — Звіт по замовленнях{period}"
    title_cell.font = Font(name="Arial", size=14, bold=True, color=HDR_BG)
    title_cell.fill = fill(LIGHT_BG)
    title_cell.alignment = center()
    ws1.row_dimensions[1].height = 28

    # Generated date
    ws1.merge_cells("A2:I2")
    gen_cell = ws1["A2"]
    gen_cell.value = f"Сформовано: {date.today().strftime('%d.%m.%Y')}"
    gen_cell.font = Font(name="Arial", size=9, italic=True, color="888888")
    gen_cell.fill = fill(LIGHT_BG)
    gen_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws1.row_dimensions[2].height = 16

    # Headers
    headers = ["#", "Клієнт", "Річ", "Послуги", "Філія",
               "Дата прийому", "Бажана дата", "Терміново", "Вартість (грн)", "Статус"]
    ws1.append([])  # row 3 empty
    ws1.row_dimensions[3].height = 6

    hdr_row = 4
    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(row=hdr_row, column=col_idx, value=h)
        cell.font = hdr_font()
        cell.fill = fill(HDR_BG)
        cell.alignment = center()
        cell.border = border()
    ws1.row_dimensions[hdr_row].height = 22

    # Data rows
    for i, order in enumerate(orders, 1):
        row_num = hdr_row + i
        is_alt = (i % 2 == 0)
        row_fill = fill(ALT_ROW) if is_alt else fill("FFFFFF")

        status = "Сплачено" if order.is_paid else ("Терміново" if order.urgent else "Очікує")
        data = [
            i,
            order.user.username,
            order.item_name,
            order.service_type,
            order.branch,
            order.order_date,
            order.desired_date,
            "Так" if order.urgent else "Ні",
            float(order.total_price),
            status,
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws1.cell(row=row_num, column=col_idx, value=val)
            cell.font = cell_font()
            cell.fill = row_fill
            cell.border = border()
            cell.alignment = left()

            if col_idx == 1:  # #
                cell.alignment = center()
                cell.font = cell_font(bold=True, color=HDR_BG)
            elif col_idx == 9:  # price
                cell.number_format = "#,##0.00"
                cell.font = cell_font(bold=True, color="6B4423")
                cell.alignment = center()
            elif col_idx == 8:  # urgent
                cell.alignment = center()
                if val == "Так":
                    cell.font = cell_font(bold=True, color="C0392B")
            elif col_idx == 10:  # status
                cell.alignment = center()
                if order.is_paid:
                    cell.font = cell_font(bold=True, color="27AE60")
                    cell.fill = fill(GREEN_BG)
                elif order.urgent:
                    cell.font = cell_font(bold=True, color="C0392B")
                    cell.fill = fill(RED_BG)
            elif col_idx in (6, 7):  # dates
                cell.number_format = "DD.MM.YYYY"
                cell.alignment = center()

        ws1.row_dimensions[row_num].height = 18

    # Totals row
    total_row = hdr_row + len(orders) + 1
    ws1.merge_cells(f"A{total_row}:H{total_row}")
    tot_label = ws1.cell(row=total_row, column=1, value="РАЗОМ")
    tot_label.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    tot_label.fill = fill(HDR_BG)
    tot_label.alignment = center()
    tot_label.border = border()

    tot_val = ws1.cell(row=total_row, column=9,
                       value=f"=SUM(I{hdr_row+1}:I{hdr_row+len(orders)})")
    tot_val.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    tot_val.fill = fill(HDR_BG)
    tot_val.number_format = "#,##0.00"
    tot_val.alignment = center()
    tot_val.border = border()

    paid_cnt = sum(1 for o in orders if o.is_paid)
    tot_status = ws1.cell(row=total_row, column=10, value=f"Сплачено: {paid_cnt}/{len(orders)}")
    tot_status.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    tot_status.fill = fill(HDR_BG)
    tot_status.alignment = center()
    tot_status.border = border()
    ws1.row_dimensions[total_row].height = 22

    # Column widths
    col_widths = [5, 28, 22, 38, 18, 14, 14, 11, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws1.freeze_panes = f"A{hdr_row+1}"

    # ═══════════════════════════════════════════════
    # SHEET 2 — Звіт по філіях
    # ═══════════════════════════════════════════════
    ws2 = wb.create_sheet("Звіт по філіях")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:F1")
    ws2["A1"].value = "Purely&Quick — Статистика по філіях"
    ws2["A1"].font = Font(name="Arial", size=14, bold=True, color=HDR_BG)
    ws2["A1"].fill = fill(LIGHT_BG)
    ws2["A1"].alignment = center()
    ws2.row_dimensions[1].height = 28

    branch_headers = ["Філія", "Замовлень", "Сплачено", "Очікує", "Виручка (грн)", "Частка виручки (%)"]
    for col_idx, h in enumerate(branch_headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = hdr_font()
        cell.fill = fill(HDR_BG)
        cell.alignment = center()
        cell.border = border()
    ws2.row_dimensions[3].height = 22

    branch_data = (
        qs.values("branch")
          .annotate(
              total=Count("id"),
              paid_c=Count("id", filter=Q(is_paid=True)),
              unpaid_c=Count("id", filter=Q(is_paid=False)),
              revenue=Sum("total_price", filter=Q(is_paid=True)),
          )
          .order_by("-revenue")
    )
    branch_map_export = {b["branch"]: b for b in branch_data}
    # Використовуємо філії з БД — щоб показати всі навіть без замовлень
    branch_list = []
    for bname in db_branches_export:
        b = branch_map_export.get(bname, {"branch": bname, "total": 0, "paid_c": 0, "unpaid_c": 0, "revenue": 0})
        b["branch"] = bname
        branch_list.append(b)
    # Додаємо філії що є в замовленнях але не в каталозі
    for bname, b in branch_map_export.items():
        if bname not in db_branches_export:
            branch_list.append(b)
    branch_list.sort(key=lambda x: float(x["revenue"] or 0), reverse=True)
    total_rev_val = sum(float(b["revenue"] or 0) for b in branch_list) or 1

    for i, b in enumerate(branch_list, 1):
        row_num = 3 + i
        rev = float(b["revenue"] or 0)
        is_alt = (i % 2 == 0)
        rf = fill(ALT_ROW) if is_alt else fill("FFFFFF")

        row_data = [
            b["branch"],
            b["total"],
            b["paid_c"],
            b["unpaid_c"],
            rev,
            f"=E{row_num}/E${3+len(branch_list)+1}",
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_num, column=col_idx, value=val)
            cell.font = cell_font()
            cell.fill = rf
            cell.border = border()
            cell.alignment = center()
            if col_idx == 1:
                cell.alignment = left()
                cell.font = cell_font(bold=True, color=HDR_BG)
            elif col_idx == 5:
                cell.number_format = "#,##0.00"
                cell.font = cell_font(bold=True, color="6B4423")
            elif col_idx == 6:
                cell.number_format = "0.0%"
                cell.font = cell_font(color="27AE60")
        ws2.row_dimensions[row_num].height = 18

    # Branch totals
    tot2 = 3 + len(branch_list) + 1
    ws2.cell(row=tot2, column=1, value="РАЗОМ").font = hdr_font()
    ws2.cell(row=tot2, column=1).fill = fill(HDR_BG)
    ws2.cell(row=tot2, column=1).alignment = center()
    ws2.cell(row=tot2, column=1).border = border()

    for col_idx in range(2, 7):
        cell = ws2.cell(row=tot2, column=col_idx,
                        value=f"=SUM({get_column_letter(col_idx)}4:{get_column_letter(col_idx)}{tot2-1})")
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = fill(HDR_BG)
        cell.border = border()
        cell.alignment = center()
        if col_idx == 5:
            cell.number_format = "#,##0.00"
        elif col_idx == 6:
            cell.number_format = "0.0%"
    ws2.row_dimensions[tot2].height = 22

    col_widths2 = [22, 14, 14, 14, 18, 20]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A4"

    # ═══════════════════════════════════════════════
    # SHEET 3 — Клієнти
    # ═══════════════════════════════════════════════
    ws3 = wb.create_sheet("Клієнти")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:E1")
    ws3["A1"].value = "Purely&Quick — База клієнтів"
    ws3["A1"].font = Font(name="Arial", size=14, bold=True, color=HDR_BG)
    ws3["A1"].fill = fill(LIGHT_BG)
    ws3["A1"].alignment = center()
    ws3.row_dimensions[1].height = 28

    cli_headers = ["Клієнт", "Всього замовлень", "Сплачено", "Загальні витрати (грн)", "Статус"]
    for col_idx, h in enumerate(cli_headers, 1):
        cell = ws3.cell(row=3, column=col_idx, value=h)
        cell.font = hdr_font()
        cell.fill = fill(HDR_BG)
        cell.alignment = center()
        cell.border = border()
    ws3.row_dimensions[3].height = 22

    clients_qs = (
        User.objects.filter(is_staff=False)
            .annotate(
                order_count=Count("order"),
                paid_count=Count("order", filter=Q(order__is_paid=True)),
                total_spent=Sum("order__total_price", filter=Q(order__is_paid=True)),
            )
            .order_by("-paid_count")
    )
    for i, c in enumerate(clients_qs, 1):
        row_num = 3 + i
        is_loyal = getattr(getattr(c, "profile", None), "paid_orders_count", 0) >= 3
        is_alt = (i % 2 == 0)
        rf = fill(ALT_ROW) if is_alt else fill("FFFFFF")

        row_data = [
            c.username,
            c.order_count,
            c.paid_count,
            float(c.total_spent or 0),
            "Постійний клієнт" if is_loyal else "Звичайний",
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_num, column=col_idx, value=val)
            cell.font = cell_font()
            cell.fill = rf
            cell.border = border()
            cell.alignment = center()
            if col_idx == 1:
                cell.alignment = left()
                cell.font = cell_font(bold=True, color=DARK)
            elif col_idx == 4:
                cell.number_format = "#,##0.00"
                cell.font = cell_font(bold=True, color="6B4423")
            elif col_idx == 5:
                if is_loyal:
                    cell.font = cell_font(bold=True, color="27AE60")
                    cell.fill = fill(GREEN_BG)
        ws3.row_dimensions[row_num].height = 18

    col_widths3 = [28, 18, 14, 22, 20]
    for i, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A4"

    # ═══════════════════════════════════════════════
    # SHEET 4 — Каталог (філії, речі, послуги з БД)
    # ═══════════════════════════════════════════════
    ws4 = wb.create_sheet("Каталог послуг")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:D1")
    ws4["A1"].value = "Purely&Quick — Каталог послуг (актуальні дані)"
    ws4["A1"].font = Font(name="Arial", size=14, bold=True, color=HDR_BG)
    ws4["A1"].fill = fill(LIGHT_BG)
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 28

    ws4.merge_cells("A2:D2")
    ws4["A2"].value = f"Дані з бази даних · Оновлено: {date.today().strftime('%d.%m.%Y')}"
    ws4["A2"].font = Font(name="Arial", size=9, italic=True, color="888888")
    ws4["A2"].fill = fill(LIGHT_BG)
    ws4["A2"].alignment = Alignment(horizontal="right", vertical="center")
    ws4.row_dimensions[2].height = 16

    current_row = 4
    for cat_name in _CATEGORY_DEFAULTS:
        obj = all_configs_export.get(cat_name)
        items_list    = obj.items    if obj else _CATEGORY_DEFAULTS[cat_name]["items"]
        services_list = obj.services if obj else _CATEGORY_DEFAULTS[cat_name]["services"]
        branches_list = obj.branches if obj else _DEFAULT_BRANCHES

        # Category header
        ws4.merge_cells(f"A{current_row}:D{current_row}")
        cat_cell = ws4.cell(row=current_row, column=1, value=f"  {cat_name}")
        cat_cell.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        cat_cell.fill = fill(HDR_BG)
        cat_cell.alignment = Alignment(horizontal="left", vertical="center")
        cat_cell.border = border()
        ws4.row_dimensions[current_row].height = 22
        current_row += 1

        # Sub-headers
        sub_headers = ["Назва речі / послуги", "Ціна (грн)", "Тип", "Філія"]
        for col_idx, h in enumerate(sub_headers, 1):
            cell = ws4.cell(row=current_row, column=col_idx, value=h)
            cell.font = Font(name="Arial", size=10, bold=True, color=HDR_BG)
            cell.fill = fill(ALT_ROW)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border()
        ws4.row_dimensions[current_row].height = 18
        current_row += 1

        # Items
        for item in items_list:
            rf = fill("FFFFFF")
            ws4.cell(row=current_row, column=1, value=item["name"]).font = cell_font()
            ws4.cell(row=current_row, column=1).fill = rf
            ws4.cell(row=current_row, column=1).border = border()
            ws4.cell(row=current_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

            price_cell = ws4.cell(row=current_row, column=2, value=item["price"])
            price_cell.font = cell_font(bold=True, color="6B4423")
            price_cell.fill = rf
            price_cell.border = border()
            price_cell.number_format = "#,##0"
            price_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws4.cell(row=current_row, column=3, value="Річ").font = cell_font(color="5a4030")
            ws4.cell(row=current_row, column=3).fill = rf
            ws4.cell(row=current_row, column=3).border = border()
            ws4.cell(row=current_row, column=3).alignment = Alignment(horizontal="center", vertical="center")

            ws4.cell(row=current_row, column=4, value="").fill = rf
            ws4.cell(row=current_row, column=4).border = border()
            ws4.row_dimensions[current_row].height = 16
            current_row += 1

        # Services
        for i, svc in enumerate(services_list):
            rf = fill(ALT_ROW) if i % 2 == 0 else fill("FFFFFF")
            ws4.cell(row=current_row, column=1, value=svc["name"]).font = cell_font()
            ws4.cell(row=current_row, column=1).fill = rf
            ws4.cell(row=current_row, column=1).border = border()
            ws4.cell(row=current_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

            price_cell = ws4.cell(row=current_row, column=2, value=svc["price"])
            price_cell.font = cell_font(bold=True, color="27AE60")
            price_cell.fill = rf
            price_cell.border = border()
            price_cell.number_format = "#,##0"
            price_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws4.cell(row=current_row, column=3, value="Послуга").font = cell_font(color="27AE60")
            ws4.cell(row=current_row, column=3).fill = rf
            ws4.cell(row=current_row, column=3).border = border()
            ws4.cell(row=current_row, column=3).alignment = Alignment(horizontal="center", vertical="center")

            ws4.cell(row=current_row, column=4, value="").fill = rf
            ws4.cell(row=current_row, column=4).border = border()
            ws4.row_dimensions[current_row].height = 16
            current_row += 1

        # Branches row
        branches_str = ", ".join(branches_list)
        ws4.merge_cells(f"A{current_row}:D{current_row}")
        br_cell = ws4.cell(row=current_row, column=1, value=f"Філії: {branches_str}")
        br_cell.font = Font(name="Arial", size=9, italic=True, color="5a8a70")
        br_cell.fill = fill("F0FAF5")
        br_cell.border = border()
        br_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws4.row_dimensions[current_row].height = 16
        current_row += 2  # gap between categories

    ws4.column_dimensions["A"].width = 35
    ws4.column_dimensions["B"].width = 14
    ws4.column_dimensions["C"].width = 12
    ws4.column_dimensions["D"].width = 14
    ws4.freeze_panes = "A4"

    # ── Response ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"PurelyQuick_zvit_{date.today().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

def register(request):
    errors = {}
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        email    = request.POST.get("email", "").strip()
        phone    = request.POST.get("phone", "").strip()
        password = request.POST.get("password", "").strip()

        if len(username.split()) < 3:
            errors["username_error"] = "Введіть ПІБ у форматі: Імʼя Прізвище По-батькові"
        elif not re.fullmatch(r"[А-ЯІЇЄҐа-яіїєґ'\-]+ [А-ЯІЇЄҐа-яіїєґ'\-]+ [А-ЯІЇЄҐа-яіїєґ'\-]+", username):
            errors["username_error"] = "ПІБ може містити лише літери українського алфавіту"

        if not re.fullmatch(r'\+380\d{9}', phone):
            errors["phone_error"] = "Телефон має бути у форматі +380XXXXXXXXX"
        elif Profile.objects.filter(phone=phone).exists():
            errors["phone_error"] = "Користувач з таким номером телефону вже існує"

        if len(password) < 8:
            errors["password_error"] = "Пароль має містити щонайменше 8 символів"

        if not errors.get("username_error") and User.objects.filter(username=username).exists():
            errors["username_error"] = "Користувач з таким ПІБ вже існує"

        if not errors:
            user = User.objects.create_user(username=username, email=email, password=password)
            profile, created = Profile.objects.get_or_create(user=user, defaults={'phone': phone})
            if not created and not profile.phone:
                profile.phone = phone
                profile.save()
            login(request, user)
            return redirect("home")

        return render(request, "main/register.html", {
            "errors": errors, "username": username, "email": email, "phone": phone,
        })

    return render(request, "main/register.html")


def login_view(request):
    error = None
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect("home")
        else:
            error = "Неправильний логін або пароль"
    return render(request, "main/login.html", {"error": error})


def logout_view(request):
    logout(request)
    return redirect("home")


@require_http_methods(["GET", "POST"])
def reviews(request):
    if request.method == "POST":
        if not request.user.is_authenticated:
            return JsonResponse({"error": "Unauthorized"}, status=401)
        try:
            data   = json.loads(request.body)
            title  = data.get("title", "").strip()
            text   = data.get("text", "").strip()
            rating = int(data.get("rating", 5))
            if not title or not text:
                return JsonResponse({"error": "Поля не заповнені"}, status=400)
            if not (1 <= rating <= 5):
                return JsonResponse({"error": "Невірний рейтинг"}, status=400)
            Review.objects.create(author=request.user, title=title, text=text, rating=rating, is_company=False)
            return JsonResponse({"success": True})
        except (json.JSONDecodeError, ValueError):
            return JsonResponse({"error": "Невірні дані"}, status=400)

    company_reviews = Review.objects.filter(is_company=True).order_by("-created_at")[:3]
    user_reviews    = Review.objects.filter(is_company=False).order_by("-created_at")[:3]
    return render(request, "main/reviews.html", {
        "company_reviews": company_reviews,
        "user_reviews":    user_reviews,
    })