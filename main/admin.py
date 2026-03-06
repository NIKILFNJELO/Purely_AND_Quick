from django.contrib import admin
from django.db.models import Sum, Count, Q
from django.urls import path
from django.shortcuts import render
from django.http import HttpResponse
from .models import Profile, Order, Review
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

admin.site.site_header = "Адмін-панель Хімчистки Purly&Quick"
admin.site.site_title = "Керування сайтом Purly&Quick"
admin.site.index_title = "Ласкаво просимо до панелі керування Purly&Quick"


def get_branch_report(date_from=None, date_to=None):
    branches = ['Нові будинки', 'Центр', 'Салтівка', 'Мотель', "П'ятихатки", 'ХТЗ']
    report = []

    for branch in branches:
        orders = Order.objects.filter(branch=branch)
        if date_from:
            orders = orders.filter(created_at__date__gte=date_from)
        if date_to:
            orders = orders.filter(created_at__date__lte=date_to)

        total = orders.count()
        urgent_count = orders.filter(urgent=True).count()
        paid_orders = orders.filter(is_paid=True)
        revenue = paid_orders.aggregate(s=Sum('total_price'))['s'] or 0

        # Постійні клієнти 
        loyal_clients = orders.filter(
            user__profile__paid_orders_count__gte=3
        ).values('user').distinct().count()

        # Сума надбавок за терміновість
        urgent_orders = orders.filter(urgent=True)
        urgent_sum = urgent_orders.aggregate(s=Sum('total_price'))['s'] or 0
        surcharge = round(float(urgent_sum) / 3, 2) 

        # Сума знижок 
        loyal_revenue = paid_orders.filter(
            user__profile__paid_orders_count__gte=3
        ).aggregate(s=Sum('total_price'))['s'] or 0
        discount_sum = round(float(loyal_revenue) * 0.03 / 0.97, 2)

        report.append({
            'branch': branch,
            'total': total,
            'urgent': urgent_count,
            'loyal_clients': loyal_clients,
            'revenue': float(revenue),
            'discount_sum': discount_sum,
            'surcharge': surcharge,
        })

    # Загальні підсумки
    totals = {
        'total': sum(r['total'] for r in report),
        'urgent': sum(r['urgent'] for r in report),
        'loyal_clients': sum(r['loyal_clients'] for r in report),
        'revenue': round(sum(r['revenue'] for r in report), 2),
        'discount_sum': round(sum(r['discount_sum'] for r in report), 2),
        'surcharge': round(sum(r['surcharge'] for r in report), 2),
    }

    return report, totals


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ('id', 'user', 'branch', 'item_name', 'service_type', 'total_price', 'is_paid', 'urgent', 'desired_date')
    list_filter = ('branch', 'is_paid', 'urgent')
    search_fields = ('user__username', 'item_name', 'branch')

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('branch-report/', self.admin_site.admin_view(self.branch_report_view), name='branch_report'),
            path('branch-report/export/', self.admin_site.admin_view(self.export_excel), name='branch_report_export'),
        ]
        return custom_urls + urls

    def branch_report_view(self, request):
        date_from = request.GET.get('date_from') or None
        date_to = request.GET.get('date_to') or None
        report, totals = get_branch_report(date_from, date_to)

        context = {
            **self.admin_site.each_context(request),
            'report': report,
            'totals': totals,
            'title': 'Звіт по філіях',
            'date_from': date_from or '',
            'date_to': date_to or '',
            'today': date.today().isoformat(),
        }
        return render(request, 'admin/branch_report.html', context)

    def export_excel(self, request):
        date_from = request.GET.get('date_from') or None
        date_to = request.GET.get('date_to') or None
        report, totals = get_branch_report(date_from, date_to)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Звіт по філіях"

        # Шапка
        ws.merge_cells('A1:H1')
        ws['A1'] = f'Звіт по філіях — Purly&Quick'
        ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
        ws['A1'].fill = PatternFill('solid', fgColor='2C3E6B')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        # Період
        ws.merge_cells('A2:H2')
        period = f"Період: {date_from or 'початок'} — {date_to or 'сьогодні'}    Дата формування: {date.today().strftime('%d.%m.%Y')}"
        ws['A2'] = period
        ws['A2'].font = Font(italic=True, color='666666')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 22

        # Заголовки
        headers = [
            'Філія', 'К-сть замовлень', 'Термінових',
            'Пост. клієнтів', 'Загальний дохід (грн)',
            'Сума знижок (грн)', 'Сума надбавок (грн)', 'Дата формування'
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='4F63FF')
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.row_dimensions[3].height = 30

        # Дані філій
        for i, r in enumerate(report, 4):
            bg = 'E8EEFF' if i % 2 == 0 else 'FFFFFF'
            values = [
                r['branch'], r['total'], r['urgent'],
                r['loyal_clients'], r['revenue'],
                r['discount_sum'], r['surcharge'],
                date.today().strftime('%d.%m.%Y')
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.fill = PatternFill('solid', fgColor=bg)
                cell.alignment = Alignment(horizontal='center')

        # Підсумки
        total_row = len(report) + 4
        ws.cell(row=total_row, column=1, value='РАЗОМ').font = Font(bold=True)
        ws.cell(row=total_row, column=1).fill = PatternFill('solid', fgColor='2C3E6B')
        ws.cell(row=total_row, column=1).font = Font(bold=True, color='FFFFFF')

        for col, key in enumerate(['total', 'urgent', 'loyal_clients', 'revenue', 'discount_sum', 'surcharge'], 2):
            cell = ws.cell(row=total_row, column=col, value=totals[key])
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='2C3E6B')
            cell.alignment = Alignment(horizontal='center')

        # Ширина
        widths = [20, 18, 14, 18, 22, 20, 20, 20]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + col)].width = w

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="branch_report.xlsx"'
        wb.save(response)
        return response


@admin.register(Profile)
class ProfileAdmin(admin.ModelAdmin):
    list_display = ('user', 'phone', 'paid_orders_count', 'loyalty_status')
    search_fields = ('user__username', 'phone')


@admin.register(Review)
class ReviewAdmin(admin.ModelAdmin):
    list_display = ('author', 'title', 'rating', 'is_company', 'created_at')